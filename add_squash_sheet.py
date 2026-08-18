"""
Adds / replaces 'Squash Junior Events' tab in the holiday spreadsheet.
AJSS data scraped directly from https://www.asiansquash.org/eventpage/ajss-event
ESF data from europeansquash.com/events | BJO from britishjunioropen.com
WSF from worldsquash.sport | US Squash from ussquash.org
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

# ── Palette ───────────────────────────────────────────────────────────────────
HDR  = PatternFill("solid", fgColor="1F4E79")
SUB  = PatternFill("solid", fgColor="2E75B6")
TH   = Side(style="thin", color="BFBFBF")
BRD  = Border(left=TH, right=TH, top=TH, bottom=TH)
CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
WHT  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
BODY = Font(name="Calibri", size=10)

LEVEL_FILLS = {
    "World Championship": PatternFill("solid", fgColor="7B3F9E"),   # deep purple
    "Asian Championship": PatternFill("solid", fgColor="CC0000"),   # red
    "Diamond":            PatternFill("solid", fgColor="B4C7E7"),   # steel blue
    "Platinum":           PatternFill("solid", fgColor="FCE4D6"),   # orange
    "Gold":               PatternFill("solid", fgColor="FFF2CC"),   # yellow
    "Silver":             PatternFill("solid", fgColor="E2EFDA"),   # green
    "Bronze":             PatternFill("solid", fgColor="D6E4F7"),   # light blue
    "National/Regional":  PatternFill("solid", fgColor="F2F2F2"),   # grey
}
ESTIMATED_FILL = PatternFill("solid", fgColor="FFE082")            # amber — estimated/unconfirmed
DARK_TEXT_LEVELS = {"World Championship", "Asian Championship"}

def ds(d): return d.strftime("%d %b %Y").lstrip("0")
def wd(d): return d.strftime("%a")

def ev(start, end, name, organiser, venue, country, level, age_groups, notes="", confirmed=True):
    return dict(start=start, end=end, name=name, organiser=organiser,
                venue=venue, country=country, level=level,
                age_groups=age_groups, notes=notes, confirmed=confirmed)

# ── Events — sorted by start date ─────────────────────────────────────────────
# ★ = confirmed from ASF website directly; others from ESF/WSF/BJO/US Squash

EVENTS = [

    # ═══════════════ 2026 ════════════════════════════════════════════════════

    ev(date(2026,  1,  2), date(2026,  1,  6),
       "British Junior Open 2026",
       "England Squash / BJO", "Birmingham, England", "England",
       "Platinum", "U11/U13/U15/U17/U19 (B&G)",
       "ESF Platinum | BJO 2027 = 2-6 Jan 2027. Source: britishjunioropen.com"),

    ev(date(2026,  1, 28), date(2026,  2,  1),
       "Sri Lanka Junior Squash Open 2026 ★",
       "Sri Lanka Squash / ASF", "Colombo, Sri Lanka", "Sri Lanka",
       "Silver", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Silver. Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026,  2,  8), date(2026,  2, 12),
       "Korea Cup Junior Open Squash Championships 2026 ★",
       "Korea Squash / ASF", "South Korea", "South Korea",
       "Silver", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Silver. Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026,  2, 11), date(2026,  2, 14),
       "Spark Education Tanglin Junior Open 2026 ★",
       "Singapore Squash / ASF", "Singapore", "Singapore",
       "Silver", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Silver. Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026,  2, 18), date(2026,  2, 22),
       "German Junior Open 2026",
       "ESF / German Squash Federation", "Hamburg, Germany", "Germany",
       "Gold", "U11/U13/U15/U17/U19 (B&G)",
       "ESF Junior Circuit Gold. Source: europeansquash.com"),

    ev(date(2026,  3, 31), date(2026,  4,  5),
       "2nd Touch 'N Go SRAFTKL International Junior Open 2026 ★",
       "SRAFTKL / ASF", "Kuala Lumpur, Malaysia", "Malaysia",
       "Gold", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Gold. Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026,  4,  7), date(2026,  4, 11),
       "Negeri Sembilan International Junior Squash Open 2026 ★",
       "Negeri Sembilan Squash / ASF", "Negeri Sembilan, Malaysia", "Malaysia",
       "Silver", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Silver. Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026,  4, 12), date(2026,  4, 16),
       "Australian Junior Open 2026 ★",
       "Squash Australia / ASF",
       "Melbourne Sports & Aquatic Centre (MSAC), Melbourne", "Australia",
       "Silver", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Silver (Zone A) trial inclusion. Also AJST Platinum & WSF Registered. "
       "Source: asiansquash.org/eventpage/ajss-event + squashaus.com.au"),

    ev(date(2026,  4, 20), date(2026,  4, 25),
       "Western India Slam 2026 ★",
       "India Squash / ASF", "India (West)", "India",
       "Gold", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Gold. Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026,  5, 20), date(2026,  5, 24),
       "33rd Asian Junior Individual Championships 2026",
       "Asian Squash Federation (ASF)",
       "Zhangyi Garden / Panzhihua, China", "China",
       "Asian Championship", "U13/U15/U17/U19 (B&G — 8 categories)",
       "8 gold medals. HKG & PAK 2 each; CHN, SGP, MAS, IND 1 each. "
       "SGP's first Asian Junior title since 1987 (Kareena Sashikumar GU13). "
       "HKG Leung Ngo San (BU13) won from 2 games down. "
       "Source: thesquashsite.com/asian-junior-champs-2026-roundup"),

    ev(date(2026,  6,  3), date(2026,  6,  7),
       "Dymon Asia Lion City Junior Open 2026 ★",
       "Singapore Squash / ASF", "Singapore", "Singapore",
       "Gold", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Gold. Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026,  6, 22), date(2026,  6, 27),
       "Eastern Slam 2026 ★",
       "India Squash / ASF", "India (East)", "India",
       "Silver", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Silver. Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026,  6, 30), date(2026,  7,  5),
       "PBA 22nd Penang Junior Open 2026 ★",
       "Penang Squash / ASF", "Penang, Malaysia", "Malaysia",
       "Gold", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Gold. Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026,  7,  2), date(2026,  7,  5),
       "Dutch Junior Open 2026",
       "ESF / Netherlands Squash", "Netherlands", "Netherlands",
       "Gold", "U11/U13/U15/U17/U19 (B&G)",
       "ESF Junior Circuit Gold. Finals confirmed 5 Jul 2026; start date estimated. Source: europeansquash.com",
       confirmed=False),

    ev(date(2026,  7,  9), date(2026,  7, 12),
       "European Junior Open 2026",
       "European Squash Federation (ESF)",
       "Sportwerk Hamburg, Germany", "Germany",
       "Platinum", "U11/U13/U15/U17/U19 (B&G)",
       "ESF Platinum. 215+ players. Second year at Hamburg. "
       "Source: europeansquash.com/event/european-junior-open-2026"),

    ev(date(2026,  7, 15), date(2026,  7, 19),
       "China Squash Junior Open 2026 ★",
       "China Squash / ASF", "China", "China",
       "Silver", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Silver. Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026,  7, 20), date(2026,  7, 25),
       "WSF World Junior Championships 2026 — Individual",
       "World Squash Federation (WSF)",
       "White Oaks Resort & Spa, Niagara-on-the-Lake, Canada", "Canada",
       "World Championship", "U19 (B&G)",
       "185 players, 24+ nations. Champions: Mohamad Zakaria (EGY — first 3-time Men's World Junior Champion), "
       "Anahat Singh (IND). First in Canada since 1984. Source: worldsquash.sport"),

    ev(date(2026,  7, 26), date(2026,  7, 31),
       "WSF World Junior Championships 2026 — Teams",
       "World Squash Federation (WSF)",
       "White Oaks Resort & Spa, Niagara-on-the-Lake, Canada", "Canada",
       "World Championship", "U19 (B&G — team format)",
       "Boys: 8 pools of 3. Girls: 4 pools of 4. Nigeria debuted in women's teams. "
       "Source: worldsquash.sport"),

    ev(date(2026,  8,  4), date(2026,  8,  9),
       "The JESSICA COMPANY Hong Kong Junior Squash Open 2026 ★",
       "Squash Association of Hong Kong, China (HKSA)",
       "Cornwall St Squash Centre & HK Squash Centre, Hong Kong", "Hong Kong",
       "Platinum", "U9/U11/U13/U15/U17/U19 (B&G — 12 categories)",
       "AJSS Platinum + WSF & PSA Satellite Tour. Rounds/QF: 4-7 Aug (Cornwall St & HKSC); "
       "SF/Final: 8-9 Aug (HKSC). PSA points: U17 & U19. 850+ players from 16 countries. "
       "Sponsored by The Jessica Company; subvented by CSTB Arts & Sport Dev Fund. "
       "Source: hksquash.org.hk + asiansquash.org"),

    ev(date(2026,  8, 12), date(2026,  8, 15),
       "19th Korea Junior Open 2026 ★",
       "Korea Squash / ASF", "South Korea", "South Korea",
       "Gold", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Gold. Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026,  8, 17), date(2026,  8, 22),
       "Southern Slam 2026 ★",
       "India Squash / ASF", "Chennai, India", "India",
       "Gold", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Gold. Entry deadline: 15 Jul 2026. Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026,  8, 19), date(2026,  8, 23),
       "37th Japan Junior Open 2026 ★",
       "Japan Squash / ASF",
       "Yokohama & Ebina, Japan", "Japan",
       "Silver", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Silver. Entry deadline: 25 Jun 2026 12:00 JST. "
       "Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026,  8, 25), date(2026,  8, 27),
       "Qatar Junior Open 2026 ★",
       "Qatar Squash / ASF", "Doha, Qatar", "Qatar",
       "Silver", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Silver. Entry deadline: 6 Aug 2026. Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026,  8, 25), date(2026,  8, 29),
       "1st Chinese Taipei Junior Squash Open 2026 ★",
       "Chinese Taipei Squash / ASF", "Taipei, Chinese Taipei", "Chinese Taipei",
       "Silver", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Silver — inaugural edition. Entry deadline: 5 Aug 2026. "
       "Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026,  8, 26), date(2026,  8, 30),
       "Macau Junior Squash Open 2026 ★",
       "Macau Squash / ASF", "Macau, China", "Macau",
       "Gold", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Gold. Entry deadline: 29 Jul 2026 12:00 Macau Time. "
       "Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026,  9, 19), date(2026,  9, 24),
       "Indian Junior Open 2026 ★",
       "Squash Rackets Federation of India / ASF", "Kolkata, India", "India",
       "Platinum", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Platinum. Entry deadline: 12 Aug 2026. "
       "Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026, 10, 10), date(2026, 10, 12),
       "US West Coast JCT 2026",
       "US Squash",
       "Redwood Shores & Redwood City, CA", "USA",
       "National/Regional", "All junior age groups",
       "US Squash Junior Championship Tour. Source: ussquash.org"),

    ev(date(2026, 10, 13), date(2026, 10, 18),
       "Milo All Star Malaysia Squash Junior Open 2026 ★",
       "Squash Rackets Association of Malaysia / ASF",
       "Kuala Lumpur, Malaysia", "Malaysia",
       "Diamond", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Diamond — highest AJSS tier. Entry deadline: 14 Sep 2026 5pm Malaysian Time. "
       "Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026, 10, 24), date(2026, 10, 28),
       "Colombo Junior Squash Open 2026 ★",
       "Sri Lanka Squash / ASF", "Ratmalana, Sri Lanka", "Sri Lanka",
       "Silver", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Silver. Entry deadline: 23 Sep 2026 23:59 Sri Lanka Time. "
       "Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026, 11, 13), date(2026, 11, 15),
       "US Mid-Atlantic JCT 2026",
       "US Squash", "Washington, D.C.", "USA",
       "National/Regional", "All junior age groups",
       "US Squash Junior Championship Tour. Source: ussquash.org"),

    ev(date(2026, 12,  1), date(2026, 12,  6),
       "REDtone 18th KL Junior Open Squash Championships 2026 ★",
       "KL Squash / ASF", "Kuala Lumpur, Malaysia", "Malaysia",
       "Platinum", "U9/U11/U13/U15/U17/U19 (B&G)",
       "AJSS Platinum. Entry deadline: 1 Oct 2026. "
       "Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026, 12,  8), date(2026, 12, 13),
       "ONCOCARE Singapore Junior Open 2026 ★",
       "Singapore Squash Rackets Association (SSRA) / ASF",
       "Singapore", "Singapore",
       "Platinum", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Platinum. Entry deadline: 1 Nov 2026. "
       "Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2026, 12, 19), date(2026, 12, 23),
       "US Junior Open 2026",
       "US Squash", "Philadelphia, PA", "USA",
       "Platinum", "U11/U13/U15/U17/U19 (B&G)",
       "One of the largest US junior squash events. Source: ussquash.org"),

    ev(date(2026, 12, 20), date(2026, 12, 26),
       "HKFC International Junior Squash Open 2026 (est.)",
       "Squash Association of Hong Kong, China (HKSA)",
       "Hong Kong Football Club & HK Squash Centre, Hong Kong", "Hong Kong",
       "Platinum", "U9/U11/U13/U15/U17/U19 (B&G — 12 categories)",
       "AJSS Platinum. NOT YET listed on ASF page — dates estimated from prior years "
       "(2025 edition: 19-24 Dec 2025). Check hksquash.org.hk for official announcement.",
       confirmed=False),

    # ═══════════════ 2027 ════════════════════════════════════════════════════

    ev(date(2027,  1,  2), date(2027,  1,  6),
       "British Junior Open 2027",
       "England Squash / BJO", "Birmingham, England", "England",
       "Platinum", "U11/U13/U15/U17/U19 (B&G)",
       "Officially confirmed 2-6 Jan 2027. 750+ players from 50+ countries. "
       "Source: britishjunioropen.com/save-the-dates-british-junior-open-2027"),

    ev(date(2027,  1, 14), date(2027,  1, 17),
       "Czech Junior Open 2027",
       "ESF / Czech Squash Federation", "Prague, Czech Republic", "Czech Republic",
       "Gold", "U11/U13/U15/U17/U19 (B&G)",
       "ESF Junior Circuit Gold. Source: europeansquash.com"),

    ev(date(2027,  1, 16), date(2027,  1, 18),
       "US Connecticut JCT 2027",
       "US Squash", "Stamford / Norwalk, CT", "USA",
       "National/Regional", "All junior age groups",
       "US Squash Junior Championship Tour. Source: ussquash.org"),

    ev(date(2027,  1, 29), date(2027,  1, 31),
       "Slovenia Junior Open 2027",
       "ESF / Squash Slovenia", "Ljubljana, Slovenia", "Slovenia",
       "Bronze", "U11/U13/U15/U17/U19 (B&G)",
       "ESF Junior Circuit Bronze. Source: europeansquash.com"),

    ev(date(2027,  2, 10), date(2027,  2, 14),
       "Sri Lanka Junior Squash Open 2027 ★",
       "Sri Lanka Squash / ASF", "Ratmalana, Sri Lanka", "Sri Lanka",
       "Silver", "U11/U13/U15/U17/U19 (B&G)",
       "AJSS Silver. Entry deadline: 4 Jan 2027. "
       "Source: asiansquash.org/eventpage/ajss-event"),

    ev(date(2027,  2, 11), date(2027,  2, 14),
       "Swiss Junior Open 2027",
       "ESF / Swiss Squash", "Langnau am Albis, Switzerland", "Switzerland",
       "Silver", "U11/U13/U15/U17/U19 (B&G)",
       "ESF Junior Circuit Silver. Source: europeansquash.com"),

    ev(date(2027,  2, 13), date(2027,  2, 15),
       "US Texas JCT 2027",
       "US Squash", "Houston, TX", "USA",
       "National/Regional", "All junior age groups",
       "US Squash Junior Championship Tour. Source: ussquash.org"),

    ev(date(2027,  2, 18), date(2027,  2, 21),
       "German Junior Open 2027",
       "ESF / German Squash Federation", "Hamburg, Germany", "Germany",
       "Gold", "U11/U13/U15/U17/U19 (B&G)",
       "ESF Junior Circuit Gold. Source: europeansquash.com"),

    ev(date(2027,  2, 19), date(2027,  2, 21),
       "US High School Championships 2027",
       "US Squash", "Philadelphia, PA", "USA",
       "National/Regional", "High School age groups",
       "US national high school squash championships. Source: ussquash.org"),

    ev(date(2027,  2, 23), date(2027,  2, 27),
       "23rd Asian Junior Team Championships 2027",
       "Asian Squash Federation (ASF)", "Thailand (venue TBC)", "Thailand",
       "Asian Championship", "U19 (B&G — team format)",
       "Confirmed on ASF 2026-27 calendar. Exact venue TBC. "
       "Source: asiansquash.org + thesquashsite.com"),

    ev(date(2027,  2, 26), date(2027,  2, 28),
       "Austrian Junior Open 2027",
       "ESF / Austrian Squash Federation", "Vienna, Austria", "Austria",
       "Silver", "U11/U13/U15/U17/U19 (B&G)",
       "ESF Junior Circuit Silver. Source: europeansquash.com"),

    ev(date(2027,  3, 12), date(2027,  3, 14),
       "US Junior Squash Championships 2027",
       "US Squash", "Philadelphia, PA", "USA",
       "National/Regional", "U11/U13/U15/U17/U19 (B&G)",
       "US national junior championships. Source: ussquash.org"),

    ev(date(2027,  3, 12), date(2027,  3, 14),
       "Croatian Junior Open 2027",
       "ESF / Croatian Squash Federation", "Zagreb, Croatia", "Croatia",
       "Bronze", "U11/U13/U15/U17/U19 (B&G)",
       "ESF Junior Circuit Bronze. Source: europeansquash.com"),

    ev(date(2027,  3, 19), date(2027,  3, 21),
       "US Junior Divisional Championships 2027",
       "US Squash", "USA (venue TBC)", "USA",
       "National/Regional", "All junior age groups",
       "US Squash regional divisional event. Source: ussquash.org"),

    ev(date(2027,  3, 19), date(2027,  3, 21),
       "Norwegian Junior Open 2027",
       "ESF / Squash Norway", "Lysaker, Norway", "Norway",
       "Bronze", "U11/U13/U15/U17/U19 (B&G)",
       "ESF Junior Circuit Bronze. Source: europeansquash.com"),

    ev(date(2027,  3, 27), date(2027,  4,  4),
       "PSA World Championships 2026-27",
       "PSA World Tour / England Squash", "London, England", "England",
       "World Championship", "Open (professional)",
       "Professional event — included for reference. Source: worldsquashchamps.com"),

    ev(date(2027,  6,  7), date(2027,  6, 13),
       "WSF World Under-23 Championships 2027",
       "World Squash Federation (WSF)", "Karachi, Pakistan", "Pakistan",
       "World Championship", "U23 (B&G)",
       "WSF global event for under-23 players. Source: worldsquash.sport"),

]

# ── Build / replace sheet ─────────────────────────────────────────────────────
COLS = [
    "Start Date", "End Date", "Day", "Tournament / Event",
    "Level / Status", "Confirmed?", "Organiser", "Venue", "Country",
    "Age Groups", "Notes / Source"
]
WIDTHS = [13, 13, 6, 46, 20, 13, 34, 42, 16, 28, 60]

def hdr_cell(ws, row, col, text, fill=HDR):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill; c.font = WHT; c.alignment = CTR; c.border = BRD

def body_cell(ws, row, col, value, fill, font, align=LFT):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill; c.font = font; c.alignment = align; c.border = BRD

PATH = r"D:\Work\claude\QL\holiday\Holiday_Calendar_2026_2027.xlsx"
wb = openpyxl.load_workbook(PATH)

if "Squash Junior Events" in wb.sheetnames:
    del wb["Squash Junior Events"]

ws = wb.create_sheet("Squash Junior Events")
ws.freeze_panes = "A3"

for i, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 30

# Title
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
t = ws.cell(row=1, column=1,
            value=("Squash Junior Open Events — 2026 & 2027  |  "
                   "★ = verified from asiansquash.org directly  |  "
                   "Sources: ASF · ESF · WSF · BJO · US Squash · HKSA · SSRA"))
t.fill = HDR
t.font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
t.alignment = CTR

# Header row
for i, label in enumerate(COLS, 1):
    hdr_cell(ws, 2, i, label, fill=SUB)

events_sorted = sorted(EVENTS, key=lambda e: e["start"])

for row_i, e in enumerate(events_sorted, start=3):
    is_confirmed = e.get("confirmed", True)
    fill = ESTIMATED_FILL if not is_confirmed else \
           LEVEL_FILLS.get(e["level"], LEVEL_FILLS["National/Regional"])
    use_white = is_confirmed and e["level"] in DARK_TEXT_LEVELS
    font = Font(color="FFFFFF", name="Calibri", size=10) if use_white else \
           Font(name="Calibri", size=10)
    vals = {
        "Start Date":         ds(e["start"]),
        "End Date":           ds(e["end"]) if e["end"] != e["start"] else "",
        "Day":                wd(e["start"]),
        "Tournament / Event": e["name"],
        "Level / Status":     e["level"],
        "Confirmed?":         "✓ Confirmed" if is_confirmed else "~ Estimated",
        "Organiser":          e["organiser"],
        "Venue":              e["venue"],
        "Country":            e["country"],
        "Age Groups":         e["age_groups"],
        "Notes / Source":     e["notes"],
    }
    for col_i, col in enumerate(COLS, 1):
        align = CTR if col in ("Start Date", "End Date", "Day", "Confirmed?") else LFT
        body_cell(ws, row_i, col_i, vals[col], fill, font, align)

# AutoFilter
last_col = get_column_letter(len(COLS))
last_row = 2 + len(events_sorted)
ws.auto_filter.ref = f"A2:{last_col}{last_row}"

# Legend
leg_row = last_row + 2
ws.merge_cells(start_row=leg_row, start_column=1, end_row=leg_row, end_column=len(COLS))
ws.cell(row=leg_row, column=1, value="COLOUR KEY — AJSS Levels: Diamond > Platinum > Gold > Silver > Bronze").fill = SUB
ws.cell(row=leg_row, column=1).font = WHT
ws.cell(row=leg_row, column=1).alignment = CTR

legend_items = [
    ("World Championship", "WSF / PSA global championship — pinnacle events"),
    ("Asian Championship",  "ASF continental championships (Asian Junior Individual & Team)"),
    ("Diamond",             "AJSS Diamond — highest AJSS tier (e.g. Milo Malaysia)"),
    ("Platinum",            "AJSS Platinum / ESF Platinum — highest open tier (HK Junior, BJO, Singapore JO, KL JO, Indian JO)"),
    ("Gold",                "AJSS Gold / ESF Gold — second tier (Korea, Japan, Penang, Lion City, Slams, Dutch JO, German JO)"),
    ("Silver",              "AJSS Silver / ESF Silver — third tier (Sri Lanka, Australia, China, Qatar, Chinese Taipei, Swiss JO)"),
    ("Bronze",              "ESF Bronze — fourth tier European circuit (Slovenia, Croatia, Norway)"),
    ("National/Regional",   "National tour stops and domestic championships (US Squash JCT / High School / Divisional)"),
    ("~ Estimated",         "AMBER = dates not yet officially confirmed — verify with the relevant federation before travel/planning"),
]
for i, (level, desc) in enumerate(legend_items, start=leg_row + 1):
    if level == "~ Estimated":
        fill = ESTIMATED_FILL
        fn = Font(name="Calibri", size=10)
    else:
        fill = LEVEL_FILLS[level]
        use_w = level in DARK_TEXT_LEVELS
        fn = Font(color="FFFFFF", name="Calibri", size=10) if use_w else Font(name="Calibri", size=10)
    c1 = ws.cell(row=i, column=1, value=level)
    c2 = ws.cell(row=i, column=2, value=desc)
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=len(COLS))
    for c in (c1, c2):
        c.fill = fill; c.font = fn; c.alignment = LFT; c.border = BRD

wb.save(PATH)
print(f"Saved: {PATH}")
print(f"  Squash Junior Events: {len(events_sorted)} events")
by_level = {}
for e in events_sorted:
    by_level[e["level"]] = by_level.get(e["level"], 0) + 1
for lv in ["World Championship","Asian Championship","Diamond","Platinum","Gold","Silver","Bronze","National/Regional"]:
    if lv in by_level:
        print(f"    {lv}: {by_level[lv]}")
