"""
Holiday Spreadsheet Builder — 2026-2027 (+ 2025-2026 reference)
Sheets: Master Calendar (filterable) | UK | US | Hong Kong | Singapore | China | Legend
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

# ── Palette ───────────────────────────────────────────────────────────────────
HDR  = PatternFill("solid", fgColor="1F4E79")
SUB  = PatternFill("solid", fgColor="2E75B6")
FILLS = {
    "UK":          PatternFill("solid", fgColor="D6E4F7"),
    "US":          PatternFill("solid", fgColor="FFF2CC"),
    "Hong Kong":   PatternFill("solid", fgColor="FCE4D6"),
    "Singapore":   PatternFill("solid", fgColor="E2EFDA"),
    "China":       PatternFill("solid", fgColor="F4CCCC"),
    "EU":          PatternFill("solid", fgColor="EAD1DC"),
    "Japan":       PatternFill("solid", fgColor="D9EAD3"),
    "Reference":   PatternFill("solid", fgColor="F2F2F2"),
}
WHT  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
BODY = Font(name="Calibri", size=10)
BOLD = Font(name="Calibri", size=10, bold=True)
TH   = Side(style="thin", color="BFBFBF")
BRD  = Border(left=TH, right=TH, top=TH, bottom=TH)
CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT  = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def ds(d): return d.strftime("%d %b %Y").lstrip("0") if hasattr(d, 'strftime') else d
def wd(d): return d.strftime("%a")
def dur(s, e): return (e - s).days + 1

def hdr_cell(ws, row, col, text, span=1, fill=HDR):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill; c.font = WHT; c.alignment = CTR; c.border = BRD
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)

def body_cell(ws, row, col, value, fill=None, align=LFT, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    if fill: c.fill = fill
    c.font = BOLD if bold else BODY
    c.alignment = align; c.border = BRD
    return c

# ── Unified data model ────────────────────────────────────────────────────────
# Each entry: dict with keys: start, end (=start if single-day), name,
#             htype, country, system, acad_year, notes
# htype: "Public Holiday" | "School Break" | "School Break - Intl"
# system: "National" | "Local/State" | "International" | "Federal" | "Market"
# country: "UK" | "US" | "Hong Kong" | "Singapore" | "China" | "EU" | "Japan"

def ph(country, start, name, notes="", acad_year=""):
    """Single-day public holiday."""
    return dict(start=start, end=start, name=name, htype="Public Holiday",
                country=country, system="National", acad_year=acad_year or str(start.year), notes=notes)

def sb(country, start, end, name, system="Local/State", notes="", acad_year=""):
    """School break period."""
    ay = acad_year or f"{start.year}/{str(end.year)[2:]}"
    return dict(start=start, end=end, name=name, htype=f"School Break - {system}",
                country=country, system=system, acad_year=ay, notes=notes)

# ══════════════════════════════════════════════════════════════════════════════
# UK
# ══════════════════════════════════════════════════════════════════════════════
UK_PH = [
    # 2025-26 reference
    ph("UK", date(2025, 8, 25), "Summer Bank Holiday"),
    ph("UK", date(2025, 12, 25), "Christmas Day"),
    ph("UK", date(2025, 12, 26), "Boxing Day"),
    ph("UK", date(2026,  1,  1), "New Year's Day"),
    ph("UK", date(2026,  4,  3), "Good Friday"),
    ph("UK", date(2026,  4,  6), "Easter Monday"),
    ph("UK", date(2026,  5,  4), "Early May Bank Holiday"),
    ph("UK", date(2026,  5, 25), "Spring Bank Holiday"),
    ph("UK", date(2026,  8, 31), "Summer Bank Holiday"),
    ph("UK", date(2026, 12, 25), "Christmas Day"),
    ph("UK", date(2026, 12, 28), "Boxing Day (substitute)"),
    ph("UK", date(2027,  1,  1), "New Year's Day"),
    ph("UK", date(2027,  3, 26), "Good Friday"),
    ph("UK", date(2027,  3, 29), "Easter Monday"),
    ph("UK", date(2027,  5,  3), "Early May Bank Holiday"),
    ph("UK", date(2027,  5, 31), "Spring Bank Holiday"),
    ph("UK", date(2027,  8, 30), "Summer Bank Holiday"),
    ph("UK", date(2027, 12, 27), "Christmas Day (substitute)"),
    ph("UK", date(2027, 12, 28), "Boxing Day (substitute)"),
]

UK_SCHOOL = [
    # 2025-26 reference (grey)
    sb("UK", date(2025, 10, 27), date(2025, 10, 31), "Autumn Half-Term 2025",  acad_year="2025/26"),
    sb("UK", date(2025, 12, 19), date(2026,  1,  4), "Christmas Holidays",      acad_year="2025/26"),
    sb("UK", date(2026,  2, 16), date(2026,  2, 20), "Spring Half-Term 2026",   acad_year="2025/26"),
    sb("UK", date(2026,  3, 30), date(2026,  4, 10), "Easter Holidays 2026",    acad_year="2025/26"),
    sb("UK", date(2026,  5, 25), date(2026,  5, 29), "May Half-Term 2026",      acad_year="2025/26"),
    sb("UK", date(2026,  7, 22), date(2026,  9,  1), "Summer Holidays 2025/26", acad_year="2025/26"),
    # 2026-27
    sb("UK", date(2026, 10, 26), date(2026, 10, 30), "Autumn Half-Term 2026",  acad_year="2026/27"),
    sb("UK", date(2026, 12, 18), date(2027,  1,  3), "Christmas Holidays",      acad_year="2026/27"),
    sb("UK", date(2027,  2, 15), date(2027,  2, 19), "Spring Half-Term 2027",   acad_year="2026/27"),
    sb("UK", date(2027,  3, 29), date(2027,  4,  9), "Easter Holidays 2027",    acad_year="2026/27"),
    sb("UK", date(2027,  5, 31), date(2027,  6,  4), "May Half-Term 2027",      acad_year="2026/27"),
    sb("UK", date(2027,  7, 22), date(2027,  9,  1), "Summer Holidays 2026/27", acad_year="2026/27"),
]

# ══════════════════════════════════════════════════════════════════════════════
# US
# ══════════════════════════════════════════════════════════════════════════════
US_PH = [
    ph("US", date(2026,  1,  1), "New Year's Day"),
    ph("US", date(2026,  1, 19), "Martin Luther King Jr. Day"),
    ph("US", date(2026,  2, 16), "Presidents' Day"),
    ph("US", date(2026,  5, 25), "Memorial Day"),
    ph("US", date(2026,  6, 19), "Juneteenth"),
    ph("US", date(2026,  7,  3), "Independence Day (observed)", notes="Sat 4 Jul → observed Fri 3 Jul"),
    ph("US", date(2026,  9,  7), "Labor Day"),
    ph("US", date(2026, 10, 12), "Columbus Day"),
    ph("US", date(2026, 11, 11), "Veterans Day"),
    ph("US", date(2026, 11, 26), "Thanksgiving Day"),
    ph("US", date(2026, 12, 25), "Christmas Day"),
    ph("US", date(2027,  1,  1), "New Year's Day"),
    ph("US", date(2027,  1, 18), "Martin Luther King Jr. Day"),
    ph("US", date(2027,  2, 15), "Presidents' Day"),
    ph("US", date(2027,  5, 31), "Memorial Day"),
    ph("US", date(2027,  6, 18), "Juneteenth (observed)", notes="Sat 19 Jun → observed Fri 18 Jun"),
    ph("US", date(2027,  7,  5), "Independence Day (observed)", notes="Sun 4 Jul → observed Mon 5 Jul"),
    ph("US", date(2027,  9,  6), "Labor Day"),
    ph("US", date(2027, 10, 11), "Columbus Day"),
    ph("US", date(2027, 11, 11), "Veterans Day"),
    ph("US", date(2027, 11, 25), "Thanksgiving Day"),
    ph("US", date(2027, 12, 24), "Christmas Day (observed)", notes="Sat 25 Dec → observed Fri 24 Dec"),
]

# ══════════════════════════════════════════════════════════════════════════════
# HONG KONG
# ══════════════════════════════════════════════════════════════════════════════
HK_PH = [
    # 2026 — CNY Day 1 = 17 Feb 2026 (Year of the Horse)
    ph("Hong Kong", date(2026,  1,  1), "New Year's Day"),
    ph("Hong Kong", date(2026,  2, 17), "Lunar New Year Day 1"),
    ph("Hong Kong", date(2026,  2, 18), "Lunar New Year Day 2"),
    ph("Hong Kong", date(2026,  2, 19), "Lunar New Year Day 3"),
    ph("Hong Kong", date(2026,  4,  3), "Good Friday"),
    ph("Hong Kong", date(2026,  4,  4), "Easter Saturday"),
    ph("Hong Kong", date(2026,  4,  6), "Easter Monday"),
    ph("Hong Kong", date(2026,  4,  7), "Ching Ming Festival (substitute)", notes="Ching Ming=5 Apr (Sun); sub falls Tue as Mon=Easter"),
    ph("Hong Kong", date(2026,  5,  1), "Labour Day"),
    ph("Hong Kong", date(2026,  5, 27), "Buddha's Birthday", notes="Approx — 4th lunar month 8th day"),
    ph("Hong Kong", date(2026,  6, 19), "Dragon Boat Festival (Tuen Ng)"),
    ph("Hong Kong", date(2026,  7,  1), "HKSAR Establishment Day"),
    ph("Hong Kong", date(2026,  9, 28), "Day following Mid-Autumn Festival", notes="Mid-Autumn=26 Sep (Sat), day after=27 Sep (Sun), sub=28 Sep (Mon)"),
    ph("Hong Kong", date(2026, 10,  1), "National Day"),
    ph("Hong Kong", date(2026, 10, 19), "Chung Yeung Festival"),
    ph("Hong Kong", date(2026, 12, 25), "Christmas Day"),
    ph("Hong Kong", date(2026, 12, 28), "First Weekday after Christmas", notes="Boxing Day 26 Dec=Sat; first weekday after=Mon 28 Dec"),
    # 2027 — CNY Day 1 = 6 Feb 2027 (Year of the Goat)
    ph("Hong Kong", date(2027,  1,  1), "New Year's Day"),
    ph("Hong Kong", date(2027,  2,  6), "Lunar New Year Day 1"),
    ph("Hong Kong", date(2027,  2,  8), "Lunar New Year Day 3"),
    ph("Hong Kong", date(2027,  2,  9), "Lunar New Year Day 4 (substitute)", notes="LNY Day 2=7 Feb (Sun); Day 4 designated as sub"),
    ph("Hong Kong", date(2027,  3, 26), "Good Friday"),
    ph("Hong Kong", date(2027,  3, 27), "Easter Saturday"),
    ph("Hong Kong", date(2027,  3, 29), "Easter Monday"),
    ph("Hong Kong", date(2027,  4,  5), "Ching Ming Festival"),
    ph("Hong Kong", date(2027,  5,  3), "Labour Day (substitute)", notes="1 May=Sat; substitute Mon 3 May"),
    ph("Hong Kong", date(2027,  5, 13), "Buddha's Birthday"),
    ph("Hong Kong", date(2027,  6,  9), "Dragon Boat Festival (Tuen Ng)"),
    ph("Hong Kong", date(2027,  7,  1), "HKSAR Establishment Day"),
    ph("Hong Kong", date(2027,  9, 28), "Day following Mid-Autumn Festival", notes="Approx — subject to HK Government gazette"),
    ph("Hong Kong", date(2027, 10,  1), "National Day"),
    ph("Hong Kong", date(2027, 10, 21), "Chung Yeung Festival"),
    ph("Hong Kong", date(2027, 12, 27), "Christmas Day (substitute)", notes="25 Dec=Sat; sub Mon 27 Dec"),
    ph("Hong Kong", date(2027, 12, 28), "First Weekday after Christmas", notes="26 Dec=Sun; sub Tue 28 Dec"),
]

HK_SCHOOL_LOCAL = [
    # EDB guideline — local state schools (Sep-Jun year, 3 terms)
    sb("Hong Kong", date(2026,  9,  2), date(2026,  9,  2),
       "Term 1 Starts (local schools)", "Local/State", acad_year="2026/27"),
    sb("Hong Kong", date(2026, 10, 17), date(2026, 10, 26),
       "October Mid-Term Break", "Local/State", acad_year="2026/27"),
    sb("Hong Kong", date(2026, 12, 21), date(2027,  1,  3),
       "Christmas & New Year Break", "Local/State", acad_year="2026/27"),
    sb("Hong Kong", date(2027,  2,  6), date(2027,  2, 16),
       "Chinese New Year Break", "Local/State", notes="Around CNY Day 1 (6 Feb)", acad_year="2026/27"),
    sb("Hong Kong", date(2027,  3, 26), date(2027,  4,  9),
       "Easter Break", "Local/State", acad_year="2026/27"),
    sb("Hong Kong", date(2027,  6, 25), date(2027,  8, 31),
       "Summer Holidays", "Local/State", acad_year="2026/27"),
]

HK_SCHOOL_INTL = [
    # Consolidated from ESF / AIS / ASHK — international school pattern
    sb("Hong Kong", date(2026,  8, 11), date(2026,  8, 11),
       "Intl Schools Start (from 11 Aug)", "International", acad_year="2026/27"),
    sb("Hong Kong", date(2026, 10, 17), date(2026, 10, 26),
       "October Half-Term", "International",
       notes="ESF: 17-26 Oct | AIS: 17-25 Oct | ASHK: 16-25 Oct", acad_year="2026/27"),
    sb("Hong Kong", date(2026, 12, 12), date(2027,  1,  4),
       "Christmas / Winter Break", "International",
       notes="ESF/CIS: 12 Dec–3 Jan | AIS: 18 Dec–4 Jan | ASHK: 19 Dec–4 Jan", acad_year="2026/27"),
    sb("Hong Kong", date(2027,  2,  5), date(2027,  2, 15),
       "Chinese New Year Break", "International",
       notes="ESF/ASHK: 5-15 Feb | AIS: 5-14 Feb | CIS: ends ~15 Feb", acad_year="2026/27"),
    sb("Hong Kong", date(2027,  3, 26), date(2027,  4, 11),
       "Spring / Easter Break", "International",
       notes="ESF: 26 Mar–11 Apr | AIS: 26 Mar–5 Apr | ASHK: 26 Mar–6 Apr", acad_year="2026/27"),
    sb("Hong Kong", date(2027,  6, 16), date(2027,  8, 10),
       "Summer Holidays", "International",
       notes="ESF ends 25 Jun | AIS ends 16 Jun | ASHK ends 17 Jun", acad_year="2026/27"),
]

# ══════════════════════════════════════════════════════════════════════════════
# SINGAPORE
# ══════════════════════════════════════════════════════════════════════════════
SG_PH = [
    # 2026 — CNY = 17 Feb, Hari Raya Puasa = 20 Mar (MOM confirmed)
    ph("Singapore", date(2026,  1,  1), "New Year's Day"),
    ph("Singapore", date(2026,  2, 17), "Chinese New Year Day 1"),
    ph("Singapore", date(2026,  2, 18), "Chinese New Year Day 2"),
    ph("Singapore", date(2026,  3, 20), "Hari Raya Puasa", notes="Subject to moon-sighting confirmation"),
    ph("Singapore", date(2026,  4,  3), "Good Friday"),
    ph("Singapore", date(2026,  5,  1), "Labour Day"),
    ph("Singapore", date(2026,  5, 27), "Hari Raya Haji", notes="Subject to moon-sighting confirmation"),
    ph("Singapore", date(2026,  6,  1), "Vesak Day (substitute)", notes="31 May=Sun; sub Mon 1 Jun"),
    ph("Singapore", date(2026,  8, 10), "National Day (substitute)", notes="9 Aug=Sun; sub Mon 10 Aug"),
    ph("Singapore", date(2026, 11,  9), "Deepavali (substitute)", notes="8 Nov=Sun; sub Mon 9 Nov"),
    ph("Singapore", date(2026, 12, 25), "Christmas Day"),
    # 2027 — CNY = 6 Feb, Hari Raya Puasa = 10 Mar (MOM gazetted Jun 2026)
    ph("Singapore", date(2027,  1,  1), "New Year's Day"),
    ph("Singapore", date(2027,  2,  6), "Chinese New Year Day 1"),
    ph("Singapore", date(2027,  2,  8), "Chinese New Year Day 2 (substitute)", notes="7 Feb=Sun; sub Mon 8 Feb"),
    ph("Singapore", date(2027,  3, 10), "Hari Raya Puasa", notes="Subject to moon-sighting confirmation"),
    ph("Singapore", date(2027,  3, 26), "Good Friday"),
    ph("Singapore", date(2027,  5,  1), "Labour Day", notes="Falls on Saturday — no automatic substitute"),
    ph("Singapore", date(2027,  5, 17), "Hari Raya Haji", notes="Subject to moon-sighting confirmation"),
    ph("Singapore", date(2027,  5, 20), "Vesak Day"),
    ph("Singapore", date(2027,  8,  9), "National Day"),
    ph("Singapore", date(2027, 12, 25), "Christmas Day", notes="Falls on Saturday — no automatic substitute"),
]

SG_SCHOOL_MOE = [
    # MOE 2026 (school year Jan-Nov)
    sb("Singapore", date(2026,  1,  2), date(2026,  1,  2),
       "MOE School Year Starts", "Local/State", acad_year="2026"),
    sb("Singapore", date(2026,  3, 19), date(2026,  4,  3),
       "Term 1 School Holidays", "Local/State", acad_year="2026",
       notes="Approx dates — confirm with MOE"),
    sb("Singapore", date(2026,  5, 30), date(2026,  6, 28),
       "June / Mid-Year Holidays", "Local/State", acad_year="2026",
       notes="Longest mid-year break"),
    sb("Singapore", date(2026,  9, 21), date(2026, 10,  2),
       "Term 3 School Holidays", "Local/State", acad_year="2026"),
    sb("Singapore", date(2026, 11, 21), date(2027,  1,  1),
       "Year-End Holidays", "Local/State", acad_year="2026",
       notes="School year ends ~20 Nov; overlaps Christmas & New Year"),
    # MOE 2027
    sb("Singapore", date(2027,  1,  4), date(2027,  1,  4),
       "MOE School Year Starts", "Local/State", acad_year="2027"),
    sb("Singapore", date(2027,  3, 13), date(2027,  3, 21),
       "Term 1 School Holidays (March)", "Local/State", acad_year="2027"),
    sb("Singapore", date(2027,  5, 29), date(2027,  6, 27),
       "June / Mid-Year Holidays", "Local/State", acad_year="2027"),
    sb("Singapore", date(2027,  9,  4), date(2027,  9, 12),
       "Term 3 School Holidays (September)", "Local/State", acad_year="2027"),
    sb("Singapore", date(2027, 11, 20), date(2027, 12, 31),
       "Year-End Holidays", "Local/State", acad_year="2027",
       notes="Provisional — MOE to confirm exact end date"),
]

SG_SCHOOL_INTL = [
    # International schools in Singapore (general Aug-Jun pattern)
    sb("Singapore", date(2026,  8, 10), date(2026,  8, 10),
       "Intl Schools Start (from ~10 Aug)", "International", acad_year="2026/27"),
    sb("Singapore", date(2026, 10, 12), date(2026, 10, 16),
       "October Half-Term / Fall Break", "International", acad_year="2026/27",
       notes="Typical 1-week break — varies by school"),
    sb("Singapore", date(2026, 12, 12), date(2027,  1,  4),
       "Christmas / Winter Break", "International", acad_year="2026/27",
       notes="Approx 3 weeks; varies by school"),
    sb("Singapore", date(2027,  2, 13), date(2027,  2, 21),
       "Chinese New Year Break", "International", acad_year="2026/27",
       notes="Approx 1 week around CNY (6 Feb 2027)"),
    sb("Singapore", date(2027,  3, 26), date(2027,  4,  9),
       "Easter / Spring Break", "International", acad_year="2026/27",
       notes="Approx 2 weeks — varies by school"),
    sb("Singapore", date(2027,  6, 14), date(2027,  8,  8),
       "Summer Holidays", "International", acad_year="2026/27",
       notes="Approx Jun-Aug; varies by school"),
]

# ══════════════════════════════════════════════════════════════════════════════
# CHINA
# ══════════════════════════════════════════════════════════════════════════════
CN_PH = [
    # 2026 — CNY Day 1 = 17 Feb (Year of the Horse)
    # Note: China holidays span multiple days; we list the full holiday period as a range
    ph("China", date(2026,  1,  1), "New Year's Day", notes="Official holiday: 1-3 Jan 2026"),
    ph("China", date(2026,  1,  2), "New Year Holiday Day 2"),
    ph("China", date(2026,  1,  3), "New Year Holiday Day 3"),
    ph("China", date(2026,  2, 15), "Spring Festival Holiday (Day 1 of 9)", notes="CNY eve; CNY Day 1=17 Feb. Makeup workdays: 14 Feb (Sat), 28 Feb (Sat)"),
    ph("China", date(2026,  2, 16), "Spring Festival Holiday Day 2"),
    ph("China", date(2026,  2, 17), "Spring Festival / Chinese New Year Day 1"),
    ph("China", date(2026,  2, 18), "Spring Festival Holiday Day 4"),
    ph("China", date(2026,  2, 19), "Spring Festival Holiday Day 5"),
    ph("China", date(2026,  2, 20), "Spring Festival Holiday Day 6"),
    ph("China", date(2026,  2, 21), "Spring Festival Holiday Day 7"),
    ph("China", date(2026,  2, 22), "Spring Festival Holiday Day 8"),
    ph("China", date(2026,  2, 23), "Spring Festival Holiday Day 9"),
    ph("China", date(2026,  4,  4), "Qingming Festival", notes="Holiday: 4-6 Apr"),
    ph("China", date(2026,  4,  5), "Qingming Festival Day 2"),
    ph("China", date(2026,  4,  6), "Qingming Festival Day 3"),
    ph("China", date(2026,  5,  1), "Labour Day / Golden Week", notes="Holiday: 1-5 May"),
    ph("China", date(2026,  5,  2), "Labour Day Holiday Day 2"),
    ph("China", date(2026,  5,  3), "Labour Day Holiday Day 3"),
    ph("China", date(2026,  5,  4), "Labour Day Holiday Day 4"),
    ph("China", date(2026,  5,  5), "Labour Day Holiday Day 5"),
    ph("China", date(2026,  6, 19), "Dragon Boat Festival", notes="Holiday: 19-21 Jun"),
    ph("China", date(2026,  6, 20), "Dragon Boat Festival Day 2"),
    ph("China", date(2026,  6, 21), "Dragon Boat Festival Day 3"),
    ph("China", date(2026,  9, 25), "Mid-Autumn Festival", notes="Holiday: 25-27 Sep"),
    ph("China", date(2026,  9, 26), "Mid-Autumn Festival Day 2"),
    ph("China", date(2026,  9, 27), "Mid-Autumn Festival Day 3"),
    ph("China", date(2026, 10,  1), "National Day / Golden Week", notes="Holiday: 1-7 Oct"),
    ph("China", date(2026, 10,  2), "National Day Holiday Day 2"),
    ph("China", date(2026, 10,  3), "National Day Holiday Day 3"),
    ph("China", date(2026, 10,  4), "National Day Holiday Day 4"),
    ph("China", date(2026, 10,  5), "National Day Holiday Day 5"),
    ph("China", date(2026, 10,  6), "National Day Holiday Day 6"),
    ph("China", date(2026, 10,  7), "National Day Holiday Day 7"),
    # 2027 — CNY Day 1 = 6 Feb (Year of the Goat). Dates provisional until State Council announcement ~Nov 2026
    ph("China", date(2027,  1,  1), "New Year's Day", notes="Holiday: 1-3 Jan 2027 (provisional)"),
    ph("China", date(2027,  1,  2), "New Year Holiday Day 2"),
    ph("China", date(2027,  1,  3), "New Year Holiday Day 3"),
    ph("China", date(2027,  2,  6), "Spring Festival / Chinese New Year Day 1", notes="Holiday approx 6-12 Feb 2027 (provisional)"),
    ph("China", date(2027,  2,  7), "Spring Festival Day 2"),
    ph("China", date(2027,  2,  8), "Spring Festival Day 3"),
    ph("China", date(2027,  2,  9), "Spring Festival Day 4"),
    ph("China", date(2027,  2, 10), "Spring Festival Day 5"),
    ph("China", date(2027,  2, 11), "Spring Festival Day 6"),
    ph("China", date(2027,  2, 12), "Spring Festival Day 7"),
    ph("China", date(2027,  4,  3), "Qingming Festival", notes="Approx 3-5 Apr 2027 (provisional)"),
    ph("China", date(2027,  4,  4), "Qingming Festival Day 2"),
    ph("China", date(2027,  4,  5), "Qingming Festival Day 3"),
    ph("China", date(2027,  5,  1), "Labour Day", notes="Approx 1-5 May 2027 (provisional)"),
    ph("China", date(2027,  5,  2), "Labour Day Holiday Day 2"),
    ph("China", date(2027,  5,  3), "Labour Day Holiday Day 3"),
    ph("China", date(2027,  5,  4), "Labour Day Holiday Day 4"),
    ph("China", date(2027,  5,  5), "Labour Day Holiday Day 5"),
    ph("China", date(2027,  6,  9), "Dragon Boat Festival", notes="Approx 9-11 Jun 2027 (provisional)"),
    ph("China", date(2027,  6, 10), "Dragon Boat Festival Day 2"),
    ph("China", date(2027,  6, 11), "Dragon Boat Festival Day 3"),
    ph("China", date(2027,  9, 27), "Mid-Autumn Festival", notes="Approx 27-29 Sep 2027 (provisional)"),
    ph("China", date(2027,  9, 28), "Mid-Autumn Festival Day 2"),
    ph("China", date(2027,  9, 29), "Mid-Autumn Festival Day 3"),
    ph("China", date(2027, 10,  1), "National Day / Golden Week", notes="1-7 Oct 2027 (provisional)"),
    ph("China", date(2027, 10,  2), "National Day Holiday Day 2"),
    ph("China", date(2027, 10,  3), "National Day Holiday Day 3"),
    ph("China", date(2027, 10,  4), "National Day Holiday Day 4"),
    ph("China", date(2027, 10,  5), "National Day Holiday Day 5"),
    ph("China", date(2027, 10,  6), "National Day Holiday Day 6"),
    ph("China", date(2027, 10,  7), "National Day Holiday Day 7"),
]

CN_SCHOOL_LOCAL = [
    # Beijing / Shanghai local school calendar (2-semester system)
    # 2025-26: winter break Jan-Mar 2026, summer break Jul-Aug 2026
    sb("China", date(2026,  1, 24), date(2026,  3,  1),
       "Winter Break 2025/26 — Beijing Local",  "Local/State",
       notes="Beijing compulsory: 24 Jan–1 Mar | Shanghai: 2 Feb–27 Feb | Northern regions earlier",
       acad_year="2025/26"),
    sb("China", date(2026,  7,  8), date(2026,  8, 31),
       "Summer Break 2025/26 — Beijing Local", "Local/State",
       notes="Beijing compulsory: 8 Jul–31 Aug | Senior high from 15 Jul",
       acad_year="2025/26"),
    # 2026-27: winter break Jan-Feb 2027 (around CNY 6 Feb), summer Jul-Aug 2027
    sb("China", date(2027,  1, 20), date(2027,  2, 28),
       "Winter Break 2026/27 — Beijing Local", "Local/State",
       notes="Provisional — set by Beijing Municipal Education Commission; aligns with CNY (6 Feb 2027)",
       acad_year="2026/27"),
    sb("China", date(2027,  7,  8), date(2027,  8, 31),
       "Summer Break 2026/27 — Beijing Local", "Local/State",
       notes="Provisional — typically Jul 8–Aug 31 for compulsory schools",
       acad_year="2026/27"),
]

CN_SCHOOL_INTL = [
    # International schools (Beijing/Shanghai) — Aug-Jun academic year
    sb("China", date(2026,  8, 17), date(2026,  8, 17),
       "Intl Schools Start (~17 Aug)", "International", acad_year="2026/27",
       notes="Dulwich/BSB/ISB/WAB start mid-to-late August"),
    sb("China", date(2026, 10,  5), date(2026, 10, 11),
       "October / Fall Break", "International", acad_year="2026/27",
       notes="Approx 1 week; aligns with China National Day Golden Week"),
    sb("China", date(2026, 12, 19), date(2027,  1,  3),
       "Christmas / Winter Break", "International", acad_year="2026/27",
       notes="Approx 2 weeks; varies by school"),
    sb("China", date(2027,  2,  6), date(2027,  2, 14),
       "Chinese New Year Break", "International", acad_year="2026/27",
       notes="Around CNY (6 Feb 2027); typically 1-2 weeks"),
    sb("China", date(2027,  3, 27), date(2027,  4,  5),
       "Spring / Easter Break", "International", acad_year="2026/27",
       notes="Around Easter (26 Mar 2027); approx 1-2 weeks"),
    sb("China", date(2027,  6, 15), date(2027,  8, 16),
       "Summer Holidays", "International", acad_year="2026/27",
       notes="Typically Jun-Aug; British schools end Jun, US schools end May/Jun"),
]

# ══════════════════════════════════════════════════════════════════════════════
# EU & Japan (for Master reference)
# ══════════════════════════════════════════════════════════════════════════════
EU_PH = [
    ph("EU", date(2026,  1,  1), "New Year's Day"),
    ph("EU", date(2026,  4,  3), "Good Friday"),
    ph("EU", date(2026,  4,  6), "Easter Monday"),
    ph("EU", date(2026,  5,  1), "Labour Day"),
    ph("EU", date(2026, 12, 25), "Christmas Day"),
    ph("EU", date(2026, 12, 26), "Second Day of Christmas"),
    ph("EU", date(2027,  1,  1), "New Year's Day"),
    ph("EU", date(2027,  3, 26), "Good Friday"),
    ph("EU", date(2027,  3, 29), "Easter Monday"),
    ph("EU", date(2027,  5,  1), "Labour Day"),
    ph("EU", date(2027, 12, 25), "Christmas Day"),
    ph("EU", date(2027, 12, 26), "Second Day of Christmas"),
]

JP_PH = [
    ph("Japan", date(2026,  1,  1), "New Year's Day"),
    ph("Japan", date(2026,  1, 12), "Coming of Age Day"),
    ph("Japan", date(2026,  2, 11), "National Foundation Day"),
    ph("Japan", date(2026,  3, 20), "Vernal Equinox Day"),
    ph("Japan", date(2026,  4, 29), "Showa Day"),
    ph("Japan", date(2026,  5,  3), "Constitution Day"),
    ph("Japan", date(2026,  5,  4), "Greenery Day"),
    ph("Japan", date(2026,  5,  5), "Children's Day"),
    ph("Japan", date(2026,  7, 20), "Marine Day"),
    ph("Japan", date(2026,  8, 11), "Mountain Day"),
    ph("Japan", date(2026,  9, 21), "Respect for the Aged Day"),
    ph("Japan", date(2026,  9, 23), "Autumnal Equinox Day"),
    ph("Japan", date(2026, 10, 12), "Sports Day"),
    ph("Japan", date(2026, 11,  3), "Culture Day"),
    ph("Japan", date(2026, 11, 23), "Labour Thanksgiving Day"),
    ph("Japan", date(2026, 12, 23), "Emperor's Birthday"),
    ph("Japan", date(2027,  1,  1), "New Year's Day"),
    ph("Japan", date(2027,  1, 11), "Coming of Age Day"),
    ph("Japan", date(2027,  2, 11), "National Foundation Day"),
    ph("Japan", date(2027,  3, 21), "Vernal Equinox Day"),
    ph("Japan", date(2027,  4, 29), "Showa Day"),
    ph("Japan", date(2027,  5,  3), "Constitution Day"),
    ph("Japan", date(2027,  5,  4), "Greenery Day"),
    ph("Japan", date(2027,  5,  5), "Children's Day"),
    ph("Japan", date(2027,  7, 19), "Marine Day"),
    ph("Japan", date(2027,  8, 11), "Mountain Day"),
    ph("Japan", date(2027,  9, 20), "Respect for the Aged Day"),
    ph("Japan", date(2027,  9, 23), "Autumnal Equinox Day"),
    ph("Japan", date(2027, 10, 11), "Sports Day"),
    ph("Japan", date(2027, 11,  3), "Culture Day"),
    ph("Japan", date(2027, 11, 23), "Labour Thanksgiving Day"),
    ph("Japan", date(2027, 12, 23), "Emperor's Birthday"),
]

ALL_DATA = (UK_PH + UK_SCHOOL +
            US_PH +
            HK_PH + HK_SCHOOL_LOCAL + HK_SCHOOL_INTL +
            SG_PH + SG_SCHOOL_MOE + SG_SCHOOL_INTL +
            CN_PH + CN_SCHOOL_LOCAL + CN_SCHOOL_INTL +
            EU_PH + JP_PH)

# ── Sheet builders ────────────────────────────────────────────────────────────

MASTER_COLS = ["Start Date","End Date","Day","Holiday / Break","Type","Country","System / Source","Acad. Year","Notes"]
MASTER_WIDTHS = [14, 14, 6, 38, 26, 14, 20, 10, 40]

COUNTRY_COLS = ["Start Date","End Date","Day","Holiday / Break","Type","System / Source","Acad. Year","Notes"]
COUNTRY_WIDTHS = [14, 14, 6, 38, 26, 20, 10, 40]

def apply_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_header_row(ws, row, cols, fill=SUB):
    for i, label in enumerate(cols, 1):
        hdr_cell(ws, row, i, label, fill=fill)

def write_entry(ws, row, entry, cols, fill):
    vals = {
        "Start Date":       ds(entry["start"]),
        "End Date":         ds(entry["end"]) if entry["end"] != entry["start"] else "",
        "Day":              wd(entry["start"]),
        "Holiday / Break":  entry["name"],
        "Type":             entry["htype"],
        "Country":          entry.get("country", ""),
        "System / Source":  entry["system"],
        "Acad. Year":       entry["acad_year"],
        "Notes":            entry["notes"],
    }
    for i, col in enumerate(cols, 1):
        align = CTR if col in ("Start Date","End Date","Day","Acad. Year") else LFT
        body_cell(ws, row, i, vals.get(col, ""), fill=fill, align=align)


def build_master(wb, data):
    ws = wb.create_sheet("Master Calendar", 0)
    ws.freeze_panes = "A3"
    apply_col_widths(ws, MASTER_WIDTHS)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16

    # Title
    span = len(MASTER_COLS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    t = ws.cell(row=1, column=1,
                value="Holiday & School Calendar 2026–2027  |  Filter by Country using column E dropdown")
    t.fill = HDR; t.font = Font(color="FFFFFF", bold=True, name="Calibri", size=12)
    t.alignment = CTR

    write_header_row(ws, 2, MASTER_COLS)

    sorted_data = sorted(data, key=lambda e: e["start"])
    for i, entry in enumerate(sorted_data, start=3):
        country = entry.get("country", "UK")
        # map country key to fill
        fill_key = country if country in FILLS else "UK"
        # Reference entries (2025/26 for UK)
        if entry.get("acad_year") == "2025/26" and country == "UK":
            fill_key = "Reference"
        fill = FILLS[fill_key]
        write_entry(ws, i, entry, MASTER_COLS, fill)

    # AutoFilter on header row
    last_col = get_column_letter(len(MASTER_COLS))
    last_row = 2 + len(sorted_data)
    ws.auto_filter.ref = f"A2:{last_col}{last_row}"
    return ws


def build_country_sheet(wb, name, ph_list, school_lists, fill_key):
    ws = wb.create_sheet(name)
    ws.freeze_panes = "A3"
    apply_col_widths(ws, COUNTRY_WIDTHS)
    ws.row_dimensions[1].height = 22

    span = len(COUNTRY_COLS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    t = ws.cell(row=1, column=1, value=f"{name}  |  Public Holidays & School Holidays 2026–2027")
    t.fill = HDR; t.font = Font(color="FFFFFF", bold=True, name="Calibri", size=12)
    t.alignment = CTR

    write_header_row(ws, 2, COUNTRY_COLS)

    all_entries = ph_list[:]
    for sl in school_lists:
        all_entries += sl
    all_entries.sort(key=lambda e: e["start"])

    fill = FILLS[fill_key]
    for i, entry in enumerate(all_entries, start=3):
        write_entry(ws, i, entry, COUNTRY_COLS, fill)

    last_col = get_column_letter(len(COUNTRY_COLS))
    last_row = 2 + len(all_entries)
    ws.auto_filter.ref = f"A2:{last_col}{last_row}"
    return ws


def build_uk_sheet(wb):
    ws = wb.create_sheet("UK")
    ws.freeze_panes = "A3"
    apply_col_widths(ws, COUNTRY_WIDTHS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COUNTRY_COLS))
    t = ws.cell(row=1, column=1,
                value="UK (England & Wales)  |  Bank Holidays + School Holidays (England state schools)  |  2025-26 Reference + 2026-27")
    t.fill = HDR; t.font = Font(color="FFFFFF", bold=True, name="Calibri", size=12)
    t.alignment = CTR

    write_header_row(ws, 2, COUNTRY_COLS)

    all_entries = sorted(UK_PH + UK_SCHOOL, key=lambda e: e["start"])
    for i, entry in enumerate(all_entries, start=3):
        fill = FILLS["Reference"] if entry.get("acad_year") == "2025/26" else FILLS["UK"]
        write_entry(ws, i, entry, COUNTRY_COLS, fill)

    last_col = get_column_letter(len(COUNTRY_COLS))
    ws.auto_filter.ref = f"A2:{last_col}{1 + len(all_entries)}"


def build_us_sheet(wb):
    ws = wb.create_sheet("US")
    ws.freeze_panes = "A3"
    apply_col_widths(ws, COUNTRY_WIDTHS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COUNTRY_COLS))
    t = ws.cell(row=1, column=1,
                value="United States  |  Federal Holidays 2026 & 2027")
    t.fill = HDR; t.font = Font(color="FFFFFF", bold=True, name="Calibri", size=12)
    t.alignment = CTR

    write_header_row(ws, 2, COUNTRY_COLS)

    all_entries = sorted(US_PH, key=lambda e: e["start"])
    fill = FILLS["US"]
    for i, entry in enumerate(all_entries, start=3):
        write_entry(ws, i, entry, COUNTRY_COLS, fill)

    last_col = get_column_letter(len(COUNTRY_COLS))
    ws.auto_filter.ref = f"A2:{last_col}{1 + len(all_entries)}"


def build_legend(wb):
    ws = wb.create_sheet("Legend & Notes")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60

    ws.merge_cells("A1:B1")
    t = ws.cell(row=1, column=1, value="Legend & Sources")
    t.fill = HDR; t.font = Font(color="FFFFFF", bold=True, name="Calibri", size=13)
    t.alignment = CTR

    sections = [
        ("COLOUR KEY", None),
        ("UK (blue)", "England & Wales bank holidays + England state school holidays"),
        ("US (yellow)", "US federal holidays (OPM official schedule)"),
        ("Hong Kong (orange)", "HK general holidays + local EDB school schedule + international school pattern"),
        ("Singapore (green)", "Singapore MOM public holidays + MOE school calendar + intl school pattern"),
        ("China (red)", "PRC national public holidays + local school breaks (Beijing/Shanghai) + intl schools"),
        ("EU (pink)", "ECB TARGET2 / Eurozone market closure days"),
        ("Japan (light green)", "Japanese national public holidays (Cabinet Office)"),
        ("Reference 2025/26 (grey)", "UK 2025-2026 school year — reference only"),
        ("SOURCES & AUTHORITY", None),
        ("UK Bank Holidays", "Gov.uk official | bankholidays2026.uk confirmed dates"),
        ("UK School Holidays", "Indicative England average; each LA/school sets own dates — always verify"),
        ("US Federal Holidays", "OPM.gov | federalpay.org official statutory schedule"),
        ("Hong Kong Public Holidays", "HK Government gazette | gov.hk/en/about/abouthk/holiday/"),
        ("HK International Schools", "ESF, AIS, ASHK, CIS 2026-27 published calendars"),
        ("Singapore Public Holidays", "MOM gazetted (Jun 2025 for 2026; Jun 2026 for 2027) | mom.gov.sg"),
        ("Singapore MOE Schools", "MOE academic calendar | moe.gov.sg/calendar"),
        ("China National Holidays", "State Council official notice | china-briefing.com"),
        ("China Local Schools", "Beijing Municipal Education Commission; dates provisional for 2026-27"),
        ("EU/Eurozone", "ECB TARGET2 settlement calendar"),
        ("Japan", "Japan Cabinet Office — national public holidays"),
        ("IMPORTANT NOTES", None),
        ("China 2027 holidays", "All 2027 China dates are PROVISIONAL — official State Council calendar released ~Nov 2026"),
        ("Lunar festival dates (HK/SG/CN)", "Hari Raya, Buddha's Birthday, Deepavali etc. subject to moon-sighting/official proclamation"),
        ("China makeup workdays", "China extends holidays by converting adjacent weekends to workdays — see Notes column"),
        ("HK/SG intl school dates", "Composite of published school calendars — verify with your specific school"),
        ("AutoFilter (Master sheet)", "Use dropdown arrows in row 2 of Master Calendar to filter by Country, Type, or Year"),
    ]

    row = 2
    for key, val in sections:
        ws.row_dimensions[row].height = 18
        if val is None:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            c = ws.cell(row=row, column=1, value=key)
            c.fill = SUB; c.font = WHT; c.alignment = CTR; c.border = BRD
        else:
            fill = None
            if "UK" in key and "Bank" not in key and "Source" not in key:
                fill = FILLS["UK"]
            elif "US" in key or "United" in key:
                fill = FILLS["US"]
            elif "Hong Kong" in key or "HK" in key:
                fill = FILLS["Hong Kong"]
            elif "Singapore" in key or "MOE" in key:
                fill = FILLS["Singapore"]
            elif "China" in key:
                fill = FILLS["China"]
            elif "EU" in key or "Eurozone" in key:
                fill = FILLS["EU"]
            elif "Japan" in key:
                fill = FILLS["Japan"]
            elif "Reference" in key or "2025/26" in key:
                fill = FILLS["Reference"]

            ka = ws.cell(row=row, column=1, value=key)
            va = ws.cell(row=row, column=2, value=val)
            for c in (ka, va):
                if fill: c.fill = fill
                c.font = BODY; c.alignment = LFT; c.border = BRD
        row += 1


# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

build_master(wb, ALL_DATA)
build_uk_sheet(wb)
build_us_sheet(wb)
build_country_sheet(wb, "Hong Kong", HK_PH, [HK_SCHOOL_LOCAL, HK_SCHOOL_INTL], "Hong Kong")
build_country_sheet(wb, "Singapore",  SG_PH, [SG_SCHOOL_MOE,  SG_SCHOOL_INTL],  "Singapore")
build_country_sheet(wb, "China",      CN_PH, [CN_SCHOOL_LOCAL, CN_SCHOOL_INTL],  "China")
build_legend(wb)

out = r"D:\Work\claude\QL\holiday\Holiday_Calendar_2026_2027.xlsx"
wb.save(out)
print(f"Saved: {out}")

# Quick summary
for sh in wb.sheetnames:
    ws = wb[sh]
    print(f"  {sh}: {ws.max_row} rows")
