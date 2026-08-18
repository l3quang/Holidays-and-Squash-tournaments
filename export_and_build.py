"""
Reads Holiday_Calendar_2026_2027.xlsx, exports each sheet to JSON,
then builds a self-contained index.html dashboard.
"""

import json, re, os
from pathlib import Path
from datetime import datetime
import openpyxl

BASE = Path(r"D:\Work\claude\QL\holiday\Holidays and Squash tournaments")
XL   = BASE / "Holiday_Calendar_2026_2027.xlsx"

# ── 1. Extract JSON ────────────────────────────────────────────────────────────

SKIP_SHEETS = {"Legend & Notes"}
HEADER_ROW  = 2   # row 1 is the title merge, row 2 is headers

wb = openpyxl.load_workbook(XL)
all_data = {}   # sheet_name -> list of dicts

for sheet_name in wb.sheetnames:
    if sheet_name in SKIP_SHEETS:
        continue
    ws = wb[sheet_name]
    headers = [ws.cell(HEADER_ROW, c).value for c in range(1, ws.max_column + 1)]
    headers = [h if h else f"col_{i}" for i, h in enumerate(headers)]

    rows = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        row = {}
        empty = True
        for c, h in enumerate(headers, 1):
            v = ws.cell(r, c).value
            if v is not None:
                empty = False
            row[h] = str(v) if v is not None else ""
        if not empty:
            rows.append(row)

    slug = re.sub(r"[^a-z0-9]+", "_", sheet_name.lower()).strip("_")
    fname = BASE / f"{slug}.json"
    with open(fname, "w", encoding="utf-8") as f:
        json.dump({"sheet": sheet_name, "headers": headers, "rows": rows}, f,
                  ensure_ascii=False, indent=2)
    print(f"  Written: {fname.name}  ({len(rows)} rows)")
    all_data[sheet_name] = {"headers": headers, "rows": rows}

# ── 2. Build HTML ──────────────────────────────────────────────────────────────

# Color maps
HOLIDAY_COLORS = {
    "UK":         "#D6E4F7",
    "US":         "#FFF2CC",
    "Hong Kong":  "#FCE4D6",
    "Singapore":  "#E2EFDA",
    "China":      "#F4CCCC",
    "EU":         "#EAD1DC",
    "Japan":      "#D9EAD3",
    "Reference":  "#F2F2F2",
}
TYPE_COLORS = {
    "Public Holiday":               "#DDEEFF",
    "School Break - Local/State":   "#D6EED6",
    "School Break - International": "#E6F4E6",
    "School Break - ISF Academy":   "#E8D5F5",
}
ISF_COLOR = "#E8D5F5"
SQUASH_COLORS = {
    "World Championship": {"bg": "#7B3F9E", "text": "#fff"},
    "Asian Championship": {"bg": "#CC0000", "text": "#fff"},
    "Diamond":            {"bg": "#B4C7E7", "text": "#000"},
    "Platinum":           {"bg": "#FCE4D6", "text": "#000"},
    "Gold":               {"bg": "#FFF2CC", "text": "#000"},
    "Silver":             {"bg": "#E2EFDA", "text": "#000"},
    "Bronze":             {"bg": "#D6E4F7", "text": "#000"},
    "National/Regional":  {"bg": "#F2F2F2", "text": "#000"},
    "Estimated":          {"bg": "#FFE082", "text": "#000"},  # amber override
}

# Tab configs: id, label, sheet name, color scheme type, filter fields
TABS = [
    {"id": "master",    "label": "All Holidays",        "sheet": "Master Calendar",    "scheme": "country"},
    {"id": "uk",        "label": "UK",                  "sheet": "UK",                 "scheme": "type",    "color": "#D6E4F7"},
    {"id": "us",        "label": "US",                  "sheet": "US",                 "scheme": "type",    "color": "#FFF2CC"},
    {"id": "hk",        "label": "Hong Kong",           "sheet": "Hong Kong",          "scheme": "type",    "color": "#FCE4D6"},
    {"id": "sg",        "label": "Singapore",           "sheet": "Singapore",          "scheme": "type",    "color": "#E2EFDA"},
    {"id": "cn",        "label": "China",               "sheet": "China",              "scheme": "type",    "color": "#F4CCCC"},
    {"id": "squash",    "label": "Squash Junior Events","sheet": "Squash Junior Events","scheme": "squash"},
]

def jstr(obj):
    return json.dumps(obj, ensure_ascii=False)

# Build tab data as JS objects
tab_js = []
for tab in TABS:
    d = all_data.get(tab["sheet"], {"headers": [], "rows": []})
    tab_js.append(f'  {jstr(tab["id"])}: {jstr(d)}')

tabs_js = "{\n" + ",\n".join(tab_js) + "\n}"

tabs_meta_js = jstr(TABS)
country_colors_js = jstr(HOLIDAY_COLORS)
type_colors_js = jstr(TYPE_COLORS)
squash_colors_js = jstr(SQUASH_COLORS)

html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Holiday & Squash Calendar 2026–2027</title>
<style>
  :root {{
    --navy:  #1F4E79;
    --blue:  #2E75B6;
    --light: #F5F8FC;
    --border:#BFBFBF;
    --text:  #1a1a1a;
    --muted: #666;
    --radius:6px;
    --shadow:0 2px 8px rgba(0,0,0,.12);
  }}
  *, *::before, *::after {{ box-sizing: border-box; margin:0; padding:0; }}
  body {{ font-family: Calibri, 'Segoe UI', sans-serif; font-size:13px;
         background:#eef3f9; color:var(--text); min-height:100vh; }}

  /* ── Header ── */
  .site-header {{
    background: linear-gradient(135deg, var(--navy) 0%, var(--blue) 100%);
    color:#fff; padding:18px 24px 0;
  }}
  .site-header h1 {{ font-size:20px; font-weight:700; letter-spacing:.3px; }}
  .site-header p  {{ font-size:12px; opacity:.8; margin-top:3px; }}

  /* ── Tabs ── */
  .tab-bar {{
    display:flex; gap:0; overflow-x:auto; margin-top:14px;
    border-bottom:none;
  }}
  .tab-btn {{
    background:rgba(255,255,255,.15); border:none; color:rgba(255,255,255,.85);
    padding:9px 18px; cursor:pointer; font-size:13px; font-family:inherit;
    border-radius:var(--radius) var(--radius) 0 0; transition:.15s;
    white-space:nowrap; border-bottom:3px solid transparent;
  }}
  .tab-btn:hover {{ background:rgba(255,255,255,.25); }}
  .tab-btn.active {{
    background:#fff; color:var(--navy); font-weight:700;
    border-bottom-color:var(--navy);
  }}

  /* ── Main ── */
  .main {{ padding:0 24px 32px; }}

  /* ── Panel ── */
  .panel {{ display:none; }}
  .panel.active {{ display:block; }}

  /* ── Toolbar ── */
  .toolbar {{
    display:flex; flex-wrap:wrap; gap:10px; align-items:center;
    padding:14px 0 10px;
  }}
  .toolbar input[type=text] {{
    flex:1; min-width:180px; max-width:300px;
    padding:7px 12px; border:1px solid var(--border); border-radius:var(--radius);
    font-size:13px; font-family:inherit; outline:none;
    transition:.15s;
  }}
  .toolbar input[type=text]:focus {{ border-color:var(--blue); box-shadow:0 0 0 2px #2E75B620; }}
  .toolbar select {{
    padding:7px 10px; border:1px solid var(--border); border-radius:var(--radius);
    font-size:13px; font-family:inherit; background:#fff; cursor:pointer;
  }}

  /* ── Multi-select dropdown ── */
  .ms-wrap {{ position:relative; }}
  .ms-btn {{
    padding:7px 10px; border:1px solid var(--border); border-radius:var(--radius);
    font-size:13px; font-family:inherit; background:#fff; cursor:pointer;
    white-space:nowrap; min-width:110px; text-align:left; line-height:1.4;
  }}
  .ms-btn:hover {{ border-color:var(--blue); }}
  .ms-panel {{
    display:none; position:absolute; top:calc(100% + 3px); left:0; z-index:200;
    background:#fff; border:1px solid var(--border); border-radius:var(--radius);
    box-shadow:var(--shadow); min-width:210px; max-height:260px; overflow-y:auto;
  }}
  .ms-wrap.open .ms-panel {{ display:block; }}
  .ms-option {{
    display:flex; align-items:center; gap:8px; padding:6px 12px;
    cursor:pointer; font-size:12.5px; user-select:none;
  }}
  .ms-option:hover {{ background:#f0f4f8; }}
  .ms-option input[type=checkbox] {{ cursor:pointer; accent-color:var(--blue); flex-shrink:0; }}
  .ms-divider {{ border:none; border-top:1px solid #E8EEF6; margin:2px 0; }}
  .isf-swatch {{
    display:inline-block; width:10px; height:10px; border-radius:2px;
    background:#E8D5F5; border:1px solid rgba(0,0,0,.15); vertical-align:middle; margin-right:3px;
  }}

  .row-count {{
    margin-left:auto; font-size:12px; color:var(--muted);
    white-space:nowrap;
  }}

  /* ── Table ── */
  .tbl-wrap {{
    overflow-x:auto; border-radius:var(--radius);
    box-shadow:var(--shadow); background:#fff;
  }}
  table {{
    border-collapse:collapse; width:100%; min-width:600px;
  }}
  thead th {{
    background:var(--blue); color:#fff; padding:9px 12px;
    text-align:left; font-size:12px; font-weight:700; white-space:nowrap;
    position:sticky; top:0; z-index:2; user-select:none; cursor:pointer;
  }}
  thead th:hover {{ background:#2563A8; }}
  thead th .sort-icon {{ margin-left:4px; opacity:.6; font-size:10px; }}
  tbody tr {{ border-bottom:1px solid #E8EEF6; transition:filter .1s; }}
  tbody tr:hover {{ filter:brightness(.95); }}
  tbody td {{
    padding:7px 12px; vertical-align:top; font-size:12.5px;
    word-break:break-word;
  }}
  .notes-col {{ max-width:320px; color:var(--muted); font-size:11.5px; }}
  .badge {{
    display:inline-block; padding:2px 8px; border-radius:999px;
    font-size:11px; font-weight:600; border:1px solid rgba(0,0,0,.08);
  }}
  .confirmed-yes {{ color:#166534; background:#dcfce7; }}
  .confirmed-no  {{ color:#92400e; background:#fef3c7; font-style:italic; }}

  /* ── Empty state ── */
  .empty-state {{
    text-align:center; padding:48px; color:var(--muted); font-size:14px;
  }}

  /* ── Squash legend ── */
  .legend {{
    display:flex; flex-wrap:wrap; gap:8px; margin-bottom:12px;
  }}
  .legend-item {{
    padding:3px 10px; border-radius:999px; font-size:11px; font-weight:600;
    border:1px solid rgba(0,0,0,.1); cursor:default;
  }}

  /* ── Stats bar ── */
  .stats-bar {{
    display:flex; gap:16px; flex-wrap:wrap; padding:10px 0 2px;
  }}
  .stat-chip {{
    background:#fff; border-radius:var(--radius); padding:6px 14px;
    box-shadow:0 1px 4px rgba(0,0,0,.08); font-size:12px;
    border-left:4px solid var(--blue);
  }}
  .stat-chip strong {{ font-size:18px; display:block; color:var(--navy); }}

  @media(max-width:600px) {{
    .site-header h1 {{ font-size:16px; }}
    .main {{ padding:0 12px 24px; }}
  }}
</style>
</head>
<body>

<div class="site-header">
  <h1>Holiday &amp; Squash Junior Calendar — 2026–2027 &nbsp;<span style="font-size:13px;font-weight:400;opacity:.85">Created for my favourite people Ari Bean Wan</span></h1>
  <p>UK · US · Hong Kong · Singapore · China · International Squash Events &nbsp;·&nbsp; Generated: {datetime.now().strftime("%d %b %Y %H:%M")}</p>
  <nav class="tab-bar" id="tabBar"></nav>
</div>

<div class="main" id="mainContent"></div>

<script>
const DATA  = {tabs_js};
const TABS  = {tabs_meta_js};
const CTRY_COLORS   = {country_colors_js};
const TYPE_COLORS   = {type_colors_js};
const SQUASH_COLORS = {squash_colors_js};
const ISF_COLOR     = {jstr(ISF_COLOR)};
const MONTH_NAMES   = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];

// ── Colour helpers ────────────────────────────────────────────────────────────
function isISF(row) {{
  return (row['Type'] || '').includes('ISF') || (row['System / Source'] || '') === 'ISF Academy';
}}
function rowStyle(tab, row) {{
  if (tab.scheme === 'squash') {{
    const est = (row['Confirmed?'] || '').includes('Estimated');
    const lvl = est ? 'Estimated' : (row['Level / Status'] || '');
    const c   = SQUASH_COLORS[lvl] || SQUASH_COLORS['National/Regional'];
    return `background:${{c.bg}};color:${{c.text}}`;
  }}
  if (isISF(row)) return `background:${{ISF_COLOR}}`;
  if (tab.scheme === 'country') {{
    const country = row['Country'] || '';
    if (country === 'UK' && (row['Acad. Year'] || '') === '2025/26')
      return 'background:#F2F2F2';
    const bg = CTRY_COLORS[country] || '#fff';
    return `background:${{bg}}`;
  }}
  if (tab.scheme === 'type') {{
    const t = row['Type'] || '';
    const bg = TYPE_COLORS[t] || (tab.color || '#fff');
    return `background:${{bg}}`;
  }}
  return '';
}}

// ── Build tab buttons ─────────────────────────────────────────────────────────
const tabBar = document.getElementById('tabBar');
const mainEl = document.getElementById('mainContent');

TABS.forEach((tab, i) => {{
  const btn = document.createElement('button');
  btn.className = 'tab-btn' + (tab.id === 'squash' ? ' active' : '');
  btn.textContent = tab.label;
  btn.dataset.tabId = tab.id;
  btn.addEventListener('click', () => activateTab(tab.id));
  tabBar.appendChild(btn);
}});

// ── Build panels ──────────────────────────────────────────────────────────────
TABS.forEach((tab, i) => {{
  const sheet = DATA[tab.id];
  if (!sheet) return;
  const panel = document.createElement('div');
  panel.className = 'panel' + (tab.id === 'squash' ? ' active' : '');
  panel.id = 'panel-' + tab.id;
  panel.innerHTML = buildPanel(tab, sheet);
  mainEl.appendChild(panel);
  initPanel(tab, sheet, panel);
}});

function buildPanel(tab, sheet) {{
  const h = sheet.headers;
  const thHTML = h.map((col, ci) =>
    `<th data-col="${{ci}}">${{esc(col)}}<span class="sort-icon">⇅</span></th>`
  ).join('');

  const filterFields = getFilterFields(tab, sheet);
  const filterHTML = filterFields.map(f =>
    `<div class="ms-wrap" data-ms-key="${{esc(f.key)}}">` +
    `<button class="ms-btn" type="button">All ${{esc(f.label)}}s ▾</button>` +
    `<div class="ms-panel"></div></div>`
  ).join('');

  return `
    <div class="toolbar">
      <input type="text" placeholder="Search…" data-search>
      ${{filterHTML}}
      <select data-month-filter title="Show from month onwards">
        <option value="">All dates</option>
      </select>
      <span class="row-count" data-rowcount></span>
    </div>
    ${{tab.id === 'squash' ? buildLegend() : ''}}
    <div class="tbl-wrap">
      <table>
        <thead><tr>${{thHTML}}</tr></thead>
        <tbody data-tbody></tbody>
      </table>
    </div>`;
}}

function buildLegend() {{
  const levels = ['World Championship','Asian Championship','Diamond','Platinum',
                  'Gold','Silver','Bronze','National/Regional'];
  const items = levels.map(lv => {{
    const c = SQUASH_COLORS[lv] || {{}};
    return `<span class="legend-item" style="background:${{c.bg}};color:${{c.text}}">${{esc(lv)}}</span>`;
  }}).join('');
  const est = `<span class="legend-item" style="background:#FFE082;color:#000">~ Estimated dates</span>`;
  return `<div class="legend">${{items}}${{est}}</div>`;
}}

function getFilterFields(tab, sheet) {{
  const fields = [];
  if (tab.scheme === 'country')
    fields.push({{key:'Country', label:'Country'}}, {{key:'Type', label:'Type'}}, {{key:'Acad. Year', label:'Year'}});
  else if (tab.scheme === 'squash')
    fields.push({{key:'Level / Status', label:'Level'}}, {{key:'Country', label:'Country'}}, {{key:'Confirmed?', label:'Confirmed'}});
  else
    fields.push({{key:'Type', label:'Type'}}, {{key:'Acad. Year', label:'Year'}});
  return fields.filter(f => sheet.headers.includes(f.key));
}}

// ── Month-year helper ─────────────────────────────────────────────────────────
function parseMonthYear(dateStr) {{
  // "17 Feb 2026" → {{label:"Feb 2026", key:2026*12+1}}
  const parts = (dateStr || '').trim().split(/\\s+/);
  if (parts.length < 3) return null;
  const mi = MONTH_NAMES.findIndex(m => m.toLowerCase() === (parts[1]||'').toLowerCase());
  const yr = parseInt(parts[2], 10);
  if (mi < 0 || isNaN(yr)) return null;
  return {{label: MONTH_NAMES[mi] + ' ' + yr, key: yr * 12 + mi}};
}}

function msBtnLabel(btn, pnl, label) {{
  const cbs = [...pnl.querySelectorAll('input.ms-cb')];
  const checked = cbs.filter(c => c.checked);
  if (!checked.length || checked.length === cbs.length)
    btn.textContent = 'All ' + label + 's ▾';
  else if (checked.length === 1)
    btn.textContent = checked[0].value + ' ▾';
  else
    btn.textContent = checked.length + ' selected ▾';
}}

function initPanel(tab, sheet, panel) {{
  const tbody    = panel.querySelector('[data-tbody]');
  const search   = panel.querySelector('[data-search]');
  const msWraps  = panel.querySelectorAll('[data-ms-key]');
  const monthSel = panel.querySelector('[data-month-filter]');
  const rcEl     = panel.querySelector('[data-rowcount]');
  const ths      = panel.querySelectorAll('thead th');

  // Populate multi-select dropdowns
  msWraps.forEach(wrap => {{
    const key   = wrap.dataset.msKey;
    const btn   = wrap.querySelector('.ms-btn');
    const pnl   = wrap.querySelector('.ms-panel');
    const label = btn.textContent.replace(/^All | ▾.*$/g, '').trim()
                    .replace(/s$/, '');
    const vals  = [...new Set(sheet.rows.map(r => r[key] || '').filter(Boolean))].sort();

    // "All" row
    const allDiv = document.createElement('div');
    allDiv.className = 'ms-option';
    const allCb = document.createElement('input');
    allCb.type = 'checkbox'; allCb.checked = true; allCb.dataset.all = '1';
    allDiv.append(allCb, Object.assign(document.createElement('span'), {{textContent: 'All'}}));
    pnl.appendChild(allDiv);
    const hr = document.createElement('hr'); hr.className = 'ms-divider';
    pnl.appendChild(hr);

    vals.forEach(v => {{
      const div = document.createElement('div'); div.className = 'ms-option';
      const cb  = document.createElement('input');
      cb.type = 'checkbox'; cb.value = v; cb.checked = true; cb.className = 'ms-cb';
      const span = document.createElement('span');
      // Add ISF swatch for ISF Academy option
      if (v === 'School Break - ISF Academy') {{
        const sw = document.createElement('span'); sw.className = 'isf-swatch';
        span.append(sw);
      }}
      span.append(document.createTextNode(v));
      div.append(cb, span);
      pnl.appendChild(div);
    }});

    btn.addEventListener('click', e => {{ e.stopPropagation(); wrap.classList.toggle('open'); }});
    pnl.addEventListener('change', e => {{
      const cb = e.target;
      if (cb.dataset.all) {{
        pnl.querySelectorAll('input.ms-cb').forEach(c => c.checked = cb.checked);
      }} else {{
        allCb.checked = [...pnl.querySelectorAll('input.ms-cb')].every(c => c.checked);
      }}
      msBtnLabel(btn, pnl, label);
      render();
    }});
  }});

  // Close dropdowns on outside click
  document.addEventListener('click', () =>
    panel.querySelectorAll('.ms-wrap.open').forEach(w => w.classList.remove('open')));

  // Populate month-year select in chronological order
  const seen = new Map();
  sheet.rows.forEach(r => {{
    const my = parseMonthYear(r['Start Date'] || '');
    if (my && !seen.has(my.key)) seen.set(my.key, my.label);
  }});
  [...seen.entries()].sort((a,b) => a[0]-b[0]).forEach(([key, label]) => {{
    const opt = document.createElement('option');
    opt.value = key; opt.textContent = 'From ' + label;
    monthSel.appendChild(opt);
  }});

  let sortCol = -1, sortAsc = true;
  let rows = [...sheet.rows];

  function getFilters() {{
    const q = (search?.value || '').toLowerCase();
    const filters = {{}};
    msWraps.forEach(wrap => {{
      const key  = wrap.dataset.msKey;
      const cbs  = [...wrap.querySelectorAll('input.ms-cb')];
      const chk  = cbs.filter(c => c.checked).map(c => c.value);
      if (chk.length < cbs.length) filters[key] = chk;
    }});
    const fromKey = monthSel && monthSel.value ? parseInt(monthSel.value, 10) : null;
    return {{q, filters, fromKey}};
  }}

  function render() {{
    const {{q, filters, fromKey}} = getFilters();
    let visible = rows.filter(row => {{
      if (q && !Object.values(row).some(v => v.toLowerCase().includes(q))) return false;
      for (const [k, allowed] of Object.entries(filters)) {{
        if (!allowed.includes(row[k])) return false;
      }}
      if (fromKey !== null) {{
        const my = parseMonthYear(row['Start Date'] || '');
        if (!my || my.key < fromKey) return false;
      }}
      return true;
    }});

    if (sortCol >= 0) {{
      const key = sheet.headers[sortCol];
      visible.sort((a, b) => {{
        const av = a[key] || '', bv = b[key] || '';
        return sortAsc ? av.localeCompare(bv) : bv.localeCompare(av);
      }});
    }}

    rcEl.textContent = `${{visible.length}} of ${{sheet.rows.length}} rows`;

    if (!visible.length) {{
      tbody.innerHTML = `<tr><td colspan="${{sheet.headers.length}}" class="empty-state">No matching rows</td></tr>`;
      return;
    }}

    tbody.innerHTML = visible.map(row => {{
      const style = rowStyle(tab, row);
      const tds = sheet.headers.map(h => {{
        let val = esc(row[h] || '');
        if (h === 'Confirmed?') {{
          const cls = val.includes('Confirmed') ? 'confirmed-yes' : 'confirmed-no';
          val = `<span class="badge ${{cls}}">${{val}}</span>`;
        }}
        const cls = h.includes('Notes') || h.includes('Source') ? ' class="notes-col"' : '';
        return `<td${{cls}}>${{val}}</td>`;
      }}).join('');
      return `<tr style="${{style}}">${{tds}}</tr>`;
    }}).join('');
  }}

  ths.forEach((th, ci) => {{
    th.addEventListener('click', () => {{
      if (sortCol === ci) sortAsc = !sortAsc; else {{ sortCol = ci; sortAsc = true; }}
      ths.forEach((t, i) => t.querySelector('.sort-icon').textContent =
        i === ci ? (sortAsc ? '↑' : '↓') : '⇅');
      render();
    }});
  }});

  search?.addEventListener('input', render);
  monthSel?.addEventListener('change', render);

  render();
}}

function activateTab(id) {{
  document.querySelectorAll('.tab-btn').forEach(b =>
    b.classList.toggle('active', b.dataset.tabId === id));
  document.querySelectorAll('.panel').forEach(p =>
    p.classList.toggle('active', p.id === 'panel-' + id));
}}

function esc(s) {{
  return String(s)
    .replace(/&/g,'&amp;').replace(/</g,'&lt;')
    .replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}}
</script>
</body>
</html>"""

out_html = BASE / "index.html"
with open(out_html, "w", encoding="utf-8") as f:
    f.write(html)
print(f"\n  Written: index.html")
print(f"\nDone. Open: {out_html}")
